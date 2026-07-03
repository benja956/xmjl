import openpyxl
import re
import sys
from datetime import datetime, timedelta

def excel_date_to_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    try:
        num = int(val)
        dt = datetime(1899, 12, 30) + timedelta(days=num)
        return dt.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        # 如果是类似 "在建" 或带有格式的文本
        return str(val).strip()

def parse_performance(text):
    if not text or not isinstance(text, str):
        return None
    text = text.strip()
    if not text:
        return None
    
    proj_name = ""
    role = ""
    area = ""
    amount = ""
    duration = ""
    record_status = ""
    
    parts = {}
    keywords = ["1、担任职位：", "2、建筑面积：", "3、合同金额：", "4、开竣工时间：", "5、四库平台："]
    
    pos = []
    for kw in keywords:
        idx = text.find(kw)
        if idx != -1:
            pos.append((idx, kw))
    
    pos.sort()
    
    if pos:
        proj_name = text[:pos[0][0]].strip()
    else:
        proj_name = text
        
    proj_name = re.sub(r'^[一二三四五六七八九十]+、', '', proj_name)
    proj_name = re.sub(r'^\d+、', '', proj_name)
    proj_name = proj_name.strip()
    
    for i in range(len(pos)):
        current_idx, current_kw = pos[i]
        next_idx = pos[i+1][0] if i+1 < len(pos) else len(text)
        val = text[current_idx + len(current_kw) : next_idx].strip()
        val = val.rstrip(',').rstrip('，').rstrip(';').rstrip('；').strip()
        parts[current_kw] = val
        
    role = parts.get("1、担任职位：", "")
    area = parts.get("2、建筑面积：", "")
    amount = parts.get("3、合同金额：", "")
    duration = parts.get("4、开竣工时间：", "")
    record_status = parts.get("5、四库平台：", "")
    
    return {
        "project_name": proj_name,
        "role": role,
        "area": area,
        "amount": amount,
        "duration": duration,
        "record_status": record_status
    }

def generate_import_sql(excel_path, sql_output_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        merged_ranges = sheet.merged_cells.ranges
        
        def get_cell_value(row, col):
            cell = sheet.cell(row=row, column=col)
            for r in merged_ranges:
                if row >= r.min_row and row <= r.max_row and col >= r.min_col and col <= r.max_col:
                    return sheet.cell(row=r.min_row, column=r.min_col).value
            return cell.value

        sql_lines = [
            "-- 清空历史数据防止重复导入",
            "DELETE FROM projects;",
            "DELETE FROM project_managers;",
            "\n"
        ]
        
        inserted_managers = set()
        
        # 数据从第 3 行开始 (第 1 行为大标题，第 2 行为列名)
        for r in range(3, sheet.max_row + 1):
            seq = get_cell_value(r, 1) # 序号
            # 如果整行都是空的，或者没有序号和姓名，说明到底了或空行
            name = get_cell_value(r, 2)
            if not name:
                continue
                
            name = str(name).strip()
            title = str(get_cell_value(r, 3) or "").strip()
            title_major = str(get_cell_value(r, 4) or "").strip()
            cert_name = str(get_cell_value(r, 5) or "").strip()
            cert_major = str(get_cell_value(r, 6) or "").strip()
            # 清洗脏数据：统一分割符为英文半角逗号，并使用正则强行切分粘连专业名
            cert_major = cert_major.replace('\n', ',').replace('，', ',').replace('、', ',').replace('；', ',').replace(';', ',')
            cert_major = re.sub(r'(工程|专业)(?=[\u4e00-\u9fa5])', r'\1,', cert_major)
            cert_major = re.sub(r',+', ',', cert_major).strip(',')
            
            safety_cert = str(get_cell_value(r, 7) or "").strip()
            performance_text = get_cell_value(r, 8)
            filing_status = str(get_cell_value(r, 9) or "").strip()
            filing_post = str(get_cell_value(r, 10) or "").strip()
            filing_start = excel_date_to_str(get_cell_value(r, 11))
            filing_end = excel_date_to_str(get_cell_value(r, 12))
            memo = str(get_cell_value(r, 13) or "").strip()
            
            # 1. 插入项目经理 (若未插入)
            if name not in inserted_managers:
                # 过滤防 SQL 注入
                esc_name = name.replace("'", "''")
                esc_title = title.replace("'", "''")
                esc_title_major = title_major.replace("'", "''")
                esc_cert_name = cert_name.replace("'", "''")
                esc_cert_major = cert_major.replace("'", "''")
                esc_safety_cert = safety_cert.replace("'", "''")
                esc_memo = memo.replace("'", "''")
                
                sql_lines.append(
                    f"INSERT INTO project_managers (name, title, title_major, cert_name, cert_major, safety_cert, memo) "
                    f"VALUES ('{esc_name}', '{esc_title}', '{esc_title_major}', '{esc_cert_name}', '{esc_cert_major}', '{esc_safety_cert}', '{esc_memo}');"
                )
                inserted_managers.add(name)
                
            # 2. 插入业绩项目
            perf = parse_performance(performance_text)
            if perf:
                esc_mname = name.replace("'", "''")
                esc_pname = perf["project_name"].replace("'", "''")
                esc_role = perf["role"].replace("'", "''")
                esc_area = perf["area"].replace("'", "''")
                esc_amount = perf["amount"].replace("'", "''")
                esc_duration = perf["duration"].replace("'", "''")
                esc_rec = perf["record_status"].replace("'", "''")
                
                esc_fstatus = filing_status.replace("'", "''")
                esc_fpost = filing_post.replace("'", "''")
                esc_fstart = filing_start.replace("'", "''")
                esc_fend = filing_end.replace("'", "''")
                
                sql_lines.append(
                    f"INSERT INTO projects (manager_name, project_name, role, area, amount, duration, record_status, filing_status, filing_post, filing_start, filing_end) "
                    f"VALUES ('{esc_mname}', '{esc_pname}', '{esc_role}', '{esc_area}', '{esc_amount}', '{esc_duration}', '{esc_rec}', '{esc_fstatus}', '{esc_fpost}', '{esc_fstart}', '{esc_fend}');"
                )
        
        # 写入 SQL 文件
        with open(sql_output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(sql_lines))
            
        print(f"Successfully generated {sql_output_path} with {len(sql_lines)} SQL statements.")
        print(f"Total Unique Managers: {len(inserted_managers)}")
        
    except Exception as e:
        print(f"Error generating SQL: {e}", file=sys.stderr)

if __name__ == "__main__":
    generate_import_sql(
        "/Users/benjamin/Documents/GitHub/xmjl/1一建业绩台账A3.xlsx",
        "/Users/benjamin/Documents/GitHub/xmjl/backend/import.sql"
    )
